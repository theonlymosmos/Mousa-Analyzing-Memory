#!/usr/bin/env python3
"""Extract Volatility3_Memory_Analysis_Reference.xlsx into data/reference.json.

Pure data extraction: no HTML. Row counts are asserted against the known-good
workbook shape and the script fails loudly on any mismatch, so a future edit
to the workbook that adds, removes, or reorders rows is caught here rather
than silently propagating into the site.
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Volatility3_Memory_Analysis_Reference.xlsx"
JSON_PATH = ROOT / "data" / "reference.json"

# sheet name -> expected data row count (excludes the 2 title rows + 1 header row)
EXPECTED_ROW_COUNTS = {
    "Usage Notes & Legend": 28,
    "Plugin Classification": 96,
    "Analysis Workflow Summary": 10,
    "Linux Plugin Classification": 59,
    "Linux Analysis Workflow Summary": 9,
    "Framework & Cross-OS Plugins": 10,
    "Triage Indicators & Gotchas": 24,
    "Volatility 2 to 3 Translation": 77,
}

DATA_START_ROW = 5  # row 1-2: title (merged), row 3: blank, row 4: header, row 5+: data


def fail(msg):
    print(f"EXTRACTION FAILED: {msg}", file=sys.stderr)
    sys.exit(1)


def cell_fill_hex(cell):
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return None
    rgb = fill.fgColor.rgb
    if not isinstance(rgb, str):
        return None
    return "#" + rgb[-6:]


def assert_row_count(sheet_name, ws):
    expected = EXPECTED_ROW_COUNTS[sheet_name]
    actual = ws.max_row - (DATA_START_ROW - 1)
    if actual != expected:
        fail(
            f"sheet '{sheet_name}' has {actual} data rows, expected {expected}. "
            "The workbook has changed shape -- update EXPECTED_ROW_COUNTS only "
            "after confirming the change is intentional."
        )


def extract_classification_sheet(ws, sheet_name):
    """Step #, Analysis Phase (Category), Plugin, Purpose, Description, Command."""
    assert_row_count(sheet_name, ws)
    rows = []
    phases = []  # ordered unique phase names with colour
    phase_colors = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        step = ws.cell(row=r, column=1).value
        phase = ws.cell(row=r, column=2).value
        plugin = ws.cell(row=r, column=3).value
        purpose = ws.cell(row=r, column=4).value
        description = ws.cell(row=r, column=5).value
        command = ws.cell(row=r, column=6).value
        color = cell_fill_hex(ws.cell(row=r, column=2))

        for field_name, val in [
            ("Step #", step),
            ("Analysis Phase (Category)", phase),
            ("Volatility 3 Plugin", plugin),
            ("Primary Purpose", purpose),
            ("Technical Description & Use Case", description),
            ("Example Command", command),
        ]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")

        if phase not in phase_colors:
            phase_colors[phase] = color
            phases.append(phase)

        rows.append(
            {
                "step": step,
                "phase": phase,
                "plugin": plugin,
                "purpose": purpose,
                "description": description,
                "command": command,
                "color": color,
            }
        )
    return rows, [{"name": p, "color": phase_colors[p]} for p in phases]


def extract_workflow_sheet(ws, sheet_name):
    """Phase #, Analysis Stage, Associated Plugins, Key Forensic Objectives."""
    assert_row_count(sheet_name, ws)
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        phase_num = ws.cell(row=r, column=1).value
        stage = ws.cell(row=r, column=2).value
        plugins = ws.cell(row=r, column=3).value
        objectives = ws.cell(row=r, column=4).value
        for field_name, val in [
            ("Phase #", phase_num),
            ("Analysis Stage", stage),
            ("Associated Plugins", plugins),
            ("Key Forensic Objectives", objectives),
        ]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")
        rows.append(
            {
                "phase_num": phase_num,
                "stage": stage,
                "plugins": [p.strip() for p in plugins.split(",")],
                "objectives": objectives,
            }
        )
    return rows


def extract_framework_sheet(ws, sheet_name):
    """Plugin, Applies To, Technical Description & Use Case, Example Command."""
    assert_row_count(sheet_name, ws)
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        plugin = ws.cell(row=r, column=1).value
        applies_to = ws.cell(row=r, column=2).value
        description = ws.cell(row=r, column=3).value
        command = ws.cell(row=r, column=4).value
        for field_name, val in [
            ("Plugin", plugin),
            ("Applies To", applies_to),
            ("Technical Description & Use Case", description),
            ("Example Command", command),
        ]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")
        rows.append(
            {
                "plugin": plugin,
                "applies_to": applies_to,
                "description": description,
                "command": command,
            }
        )
    return rows


def extract_triage_sheet(ws, sheet_name):
    """Area, What To Look For, Why It Matters / Pitfall."""
    assert_row_count(sheet_name, ws)
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        area = ws.cell(row=r, column=1).value
        look_for = ws.cell(row=r, column=2).value
        why = ws.cell(row=r, column=3).value
        for field_name, val in [
            ("Area", area),
            ("What To Look For", look_for),
            ("Why It Matters / Pitfall", why),
        ]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")
        rows.append({"area": area, "look_for": look_for, "why": why})
    return rows


def extract_notes_sheet(ws, sheet_name):
    """Topic, Flag / Syntax, Notes -- plus a trailing phase-colour legend block."""
    assert_row_count(sheet_name, ws)
    notes = []
    legend = []
    r = DATA_START_ROW
    while r <= ws.max_row:
        topic = ws.cell(row=r, column=1).value
        flag = ws.cell(row=r, column=2).value
        note = ws.cell(row=r, column=3).value
        if topic is None and flag is None and note is None:
            r += 1
            continue
        if topic == "Legend — phase colour bands":
            r += 1
            continue
        if isinstance(topic, str) and topic.startswith("Phase "):
            color = cell_fill_hex(ws.cell(row=r, column=1))
            legend.append({"phase": topic, "description": flag, "color": color})
            r += 1
            continue
        for field_name, val in [("Topic", topic), ("Flag / Syntax", flag), ("Notes", note)]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")
        notes.append({"topic": topic, "flag": flag, "notes": note})
        r += 1
    if len(legend) != 10:
        fail(f"sheet '{sheet_name}': expected 10 legend phase rows, found {len(legend)}")
    return notes, legend


def extract_translation_sheet(ws, sheet_name):
    """Section-headed rows: Volatility 2 Command, Volatility 3 Equivalent, Notes."""
    assert_row_count(sheet_name, ws)
    sections = []
    current = None
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v2 = ws.cell(row=r, column=1).value
        v3 = ws.cell(row=r, column=2).value
        notes = ws.cell(row=r, column=3).value
        is_section_header = v3 is None and notes is None
        if is_section_header:
            if v2 is None or (isinstance(v2, str) and v2.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty section header")
            current = {"section": v2, "rows": []}
            sections.append(current)
            continue
        if current is None:
            fail(f"sheet '{sheet_name}' row {r}: data row before any section header")
        for field_name, val in [
            ("Volatility 2 Command", v2),
            ("Volatility 3 Equivalent", v3),
            ("Notes & Differences", notes),
        ]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"sheet '{sheet_name}' row {r}: empty '{field_name}'")
        current["rows"].append({"v2": v2, "v3": v3, "notes": notes})
    return sections


def main():
    if not XLSX_PATH.exists():
        fail(f"workbook not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    expected_sheets = list(EXPECTED_ROW_COUNTS.keys())
    if wb.sheetnames != expected_sheets:
        fail(f"sheet list mismatch.\nexpected: {expected_sheets}\nactual:   {wb.sheetnames}")

    windows_rows, windows_phases = extract_classification_sheet(
        wb["Plugin Classification"], "Plugin Classification"
    )
    linux_rows, linux_phases = extract_classification_sheet(
        wb["Linux Plugin Classification"], "Linux Plugin Classification"
    )
    windows_workflow = extract_workflow_sheet(
        wb["Analysis Workflow Summary"], "Analysis Workflow Summary"
    )
    linux_workflow = extract_workflow_sheet(
        wb["Linux Analysis Workflow Summary"], "Linux Analysis Workflow Summary"
    )
    framework_rows = extract_framework_sheet(
        wb["Framework & Cross-OS Plugins"], "Framework & Cross-OS Plugins"
    )
    triage_rows = extract_triage_sheet(wb["Triage Indicators & Gotchas"], "Triage Indicators & Gotchas")
    notes_rows, legend_rows = extract_notes_sheet(wb["Usage Notes & Legend"], "Usage Notes & Legend")
    translation_sections = extract_translation_sheet(
        wb["Volatility 2 to 3 Translation"], "Volatility 2 to 3 Translation"
    )

    if len(windows_phases) != 10:
        fail(f"expected 10 Windows phases, found {len(windows_phases)}")
    if len(linux_phases) != 9:
        fail(f"expected 9 Linux phases, found {len(linux_phases)}")

    data = {
        "meta": {
            "source_workbook": XLSX_PATH.name,
            "verified_version": "2.28",
        },
        "windows": {"phases": windows_phases, "rows": windows_rows},
        "linux": {"phases": linux_phases, "rows": linux_rows},
        "windows_workflow": windows_workflow,
        "linux_workflow": linux_workflow,
        "framework": framework_rows,
        "triage": triage_rows,
        "notes": notes_rows,
        "legend": legend_rows,
        "translation": translation_sections,
    }

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    JSON_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    total_translation_rows = sum(len(s["rows"]) for s in translation_sections)
    print("Extraction OK:")
    print(f"  windows plugins:        {len(windows_rows)}")
    print(f"  linux plugins:          {len(linux_rows)}")
    print(f"  windows workflow rows:  {len(windows_workflow)}")
    print(f"  linux workflow rows:    {len(linux_workflow)}")
    print(f"  framework rows:         {len(framework_rows)}")
    print(f"  triage rows:            {len(triage_rows)}")
    print(f"  notes rows:             {len(notes_rows)}")
    print(f"  legend rows:            {len(legend_rows)}")
    print(f"  translation sections:   {len(translation_sections)} ({total_translation_rows} data rows)")
    print(f"  wrote {JSON_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
